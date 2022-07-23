import re
from dataclasses import dataclass
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict

extract_pattern = re.compile(r"^DOMAIN: ([\w\s]+)\s\((\w+)\)$")
nist_pattern = re.compile(r"^NIST.*\s(\d+\.\d+\.\d+)$")


@dataclass
class CmmcMapping:
    key: str
    name: str
    description: str
    references: List[str]
    nist_key: str


def parse(file_path: str) -> List[CmmcMapping]:
    workbook = load_workbook(file_path)
    sheets: List[Worksheet] = [workbook[x] for x in workbook.sheetnames if len(x) == 2]
    all_entities = []
    for sheet in sheets:
        header = sheet.cell(1, 1).value
        match = extract_pattern.findall(header)
        assert match
        domain_full, domain_short = match[0]

        raw_items = extract_column(sheet, 1) + extract_column(sheet, 2) + extract_column(sheet, 3)
        all_entities += [parse_raw_item(domain_full, domain_short, t) for t in raw_items]

    return all_entities


def parse_raw_item(domain_full: str, domain_short: str, text: str) -> CmmcMapping:
    _, key, name, description, *references = text.split("\n")
    result = {"key": key.strip(), "name": name.strip(), "description": description.strip(), "references": []}
    for x in references:
        x = x.strip("•").strip()
        if not x:
            continue
        result['references'].append(x)
        nist_match = nist_pattern.findall(x)
        if nist_match:
            result["nist_key"] = nist_match[0]

    assert "nist_key" in result
    return CmmcMapping(**result)


def extract_column(sheet: Worksheet, column: int) -> List[str]:
    row = 4
    results = []
    while sheet.cell(row, column).value:
        contents = sheet.cell(row, column).value
        results.append(contents)
        row += 1

    return results
