import re
from dataclasses import dataclass
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict


@dataclass
class NistMapping:
    nist_key: str
    csf_key: str
    description: str


def parse(file_path: str) -> List[NistMapping]:
    workbook: Workbook = load_workbook(file_path)
    sheet: Worksheet = workbook[workbook.sheetnames[0]]

    results = []
    row = 5
    while sheet.cell(row, 2).value:
        csf = sheet.cell(row, 2).value
        nist = sheet.cell(row, 3).value
        description = sheet.cell(row, 4).value
        row += 1

        if nist:
            results.append(NistMapping(nist, csf, description))

    return results
